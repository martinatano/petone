"""
extraer_fotos_excel.py
──────────────────────
Extrae imágenes del formato MODERNO de Excel (cell images / richValue),
que es lo que usa Microsoft 365 / Excel 2021+.

Estructura del Excel:
  id | nombre | u por caja | descripcion | precio | categoria | foto1 | foto2 | foto3

USO:
    python extraer_fotos_excel.py productos.xlsx

REQUISITOS:
    pip install openpyxl Pillow
"""

import sys, os, zipfile, io, csv, re
from pathlib import Path

try:
    from PIL import Image
except ImportError:
    os.system(f"{sys.executable} -m pip install Pillow --quiet")
    from PIL import Image

try:
    import openpyxl
except ImportError:
    os.system(f"{sys.executable} -m pip install openpyxl --quiet")
    import openpyxl

# ── CONFIGURACIÓN ────────────────────────────────────────────────────
COL_ID          = "id"
COL_NOMBRE      = "nombre"
COL_UNIDADES    = "u por caja"
COL_DESCRIPCION = "descripcion"
COL_PRECIO      = "precio"
COL_CATEGORIA   = "categoria"
# Las columnas foto1/foto2/foto3 tienen imágenes incrustadas (no texto)

IMG_OUTPUT_DIR  = Path("imagenes")
CSV_OUTPUT      = Path("productos.csv")
IMG_MAX_SIZE    = (1200, 1200)
IMG_QUALITY     = 82
IMG_MAX_KB      = 150
# ─────────────────────────────────────────────────────────────────────

def limpiar_precio(valor):
    if valor is None: return ""
    if isinstance(valor, (int, float)): return str(int(valor))
    s = re.sub(r'[^\d]', '', str(valor))
    return s if s else ""

def build_image_map(xlsx_path):
    """
    Construye el mapa: (fila, col_letra) → bytes_imagen
    usando el sistema richValue de Excel moderno.
    """
    with zipfile.ZipFile(xlsx_path) as z:
        all_files = z.namelist()

        # 1. Cargar todas las imágenes de xl/media/
        media = {}
        for name in all_files:
            if name.startswith("xl/media/"):
                media[Path(name).name] = z.read(name)

        # 2. Leer richValueRel.xml → lista ordenada de rIds
        if 'xl/richData/richValueRel.xml' not in all_files:
            return {}, media
        rv_xml = z.read('xl/richData/richValueRel.xml').decode('utf-8')
        rids_ordered = re.findall(r'<rel r:id="(rId\d+)"', rv_xml)

        # 3. Leer .rels → mapa rId → nombre de archivo imagen
        rels_path = 'xl/richData/_rels/richValueRel.xml.rels'
        if rels_path not in all_files:
            return {}, media
        rels_xml = z.read(rels_path).decode('utf-8')
        rel_map = dict(re.findall(
            r'Id="(rId\d+)"[^>]*Target="(?:\.\./)?media/([^"]+)"', rels_xml
        ))

        # 4. Leer sheet XML → mapa (fila, col) → vm_index
        sheet_xml = z.read('xl/worksheets/sheet1.xml').decode('utf-8')
        rows_content = re.findall(r'<row r="(\d+)"[^>]*>(.*?)</row>', sheet_xml, re.DOTALL)

        cell_to_img = {}
        for row_num_str, row_content in rows_content:
            row_num = int(row_num_str)
            # buscar todas las celdas con vm= en esta fila
            cells_with_vm = re.findall(r'<c r="([A-Z]+\d+)"[^>]*vm="(\d+)"', row_content)
            for cell_ref, vm_str in cells_with_vm:
                vm = int(vm_str)
                col_letter = re.match(r'([A-Z]+)', cell_ref).group(1)
                # vm es 1-based
                if vm - 1 < len(rids_ordered):
                    rid = rids_ordered[vm - 1]
                    img_name = rel_map.get(rid)
                    if img_name and img_name in media:
                        cell_to_img[(row_num, col_letter)] = media[img_name]

    return cell_to_img, media

def comprimir(img_bytes, ruta_salida):
    img = Image.open(io.BytesIO(img_bytes))
    if img.mode in ("RGBA", "P", "LA"):
        fondo = Image.new("RGB", img.size, (255, 255, 255))
        if img.mode == "P": img = img.convert("RGBA")
        fondo.paste(img, mask=img.split()[-1] if img.mode in ("RGBA","LA") else None)
        img = fondo
    elif img.mode != "RGB":
        img = img.convert("RGB")
    img.thumbnail(IMG_MAX_SIZE, Image.LANCZOS)
    quality = IMG_QUALITY
    while True:
        buf = io.BytesIO()
        img.save(buf, format="JPEG", quality=quality, optimize=True)
        if buf.tell() / 1024 <= IMG_MAX_KB or quality <= 55: break
        quality -= 5
    with open(ruta_salida, "wb") as f:
        f.write(buf.getvalue())
    return buf.tell() / 1024

def get_col_letter(ws, header_name):
    """Devuelve la letra de columna (ej: 'G') para un encabezado dado."""
    for cell in ws[1]:
        if cell.value and str(cell.value).strip().lower() == header_name.lower():
            # convertir número de columna a letra
            from openpyxl.utils import get_column_letter
            return get_column_letter(cell.column)
    return None

def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)

    xlsx_path = Path(sys.argv[1])
    if not xlsx_path.exists():
        print(f"✗ Archivo no encontrado: {xlsx_path}")
        sys.exit(1)

    print(f"\n📂 Procesando {xlsx_path.name}...")

    # construir mapa de imágenes
    cell_to_img, media = build_image_map(xlsx_path)
    print(f"   Imágenes en archivo : {len(media)}")
    print(f"   Imágenes mapeadas   : {len(cell_to_img)}")

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    headers = [c.value for c in ws[1] if c.value]
    print(f"   Columnas            : {headers}")

    # columnas de datos (índice numérico)
    def col_idx(name):
        for cell in ws[1]:
            if cell.value and str(cell.value).strip().lower() == name.lower():
                return cell.column
        return None

    from openpyxl.utils import get_column_letter

    cols = {
        "id":   col_idx(COL_ID),
        "nom":  col_idx(COL_NOMBRE),
        "uds":  col_idx(COL_UNIDADES),
        "desc": col_idx(COL_DESCRIPCION),
        "prec": col_idx(COL_PRECIO),
        "cat":  col_idx(COL_CATEGORIA),
    }

    # columnas de fotos (pueden tener imágenes richValue)
    foto_cols = []
    for fname in ["foto1", "foto2", "foto3"]:
        ci = col_idx(fname)
        if ci:
            foto_cols.append(get_column_letter(ci))

    print(f"   Columnas de fotos   : {foto_cols}")
    if not cols["id"]:
        print(f"\n✗ No se encontró columna '{COL_ID}'. Ajustá COL_* al inicio del script.")
        sys.exit(1)

    IMG_OUTPUT_DIR.mkdir(exist_ok=True)
    rows_csv = []
    errores  = []
    sin_foto = []
    last_img_bytes = {}  # col_letter → última imagen vista (para celdas combinadas)

    print(f"\n🔄 Exportando {ws.max_row - 1} productos...\n")

    for row_num in range(2, ws.max_row + 1):
        id_val = ws.cell(row_num, cols["id"]).value
        if not id_val: continue

        id_str   = str(id_val).strip()
        nom_val  = ws.cell(row_num, cols["nom"]).value  if cols["nom"]  else ""
        uds_val  = ws.cell(row_num, cols["uds"]).value  if cols["uds"]  else ""
        desc_val = ws.cell(row_num, cols["desc"]).value if cols["desc"] else ""
        prec_val = ws.cell(row_num, cols["prec"]).value if cols["prec"] else ""
        cat_val  = ws.cell(row_num, cols["cat"]).value  if cols["cat"]  else ""

        precio = limpiar_precio(prec_val)

        # descripción limpia (sin mezclar unidades)
        desc_str = str(desc_val or "").strip()
        desc_str = "" if desc_str in ("-", "") else desc_str
        uds_str  = str(uds_val or "").strip()
        uds_str  = "" if uds_str in ("-", "0", "") else uds_str

        # extraer fotos — si la celda no tiene imagen propia, reusar la
        # última vista en esa columna (caso celda combinada entre filas)
        nombres_foto = []
        for i, col_letter in enumerate(foto_cols, start=1):
            img_bytes = cell_to_img.get((row_num, col_letter))
            if img_bytes:
                # imagen nueva: guardarla como última vista
                last_img_bytes[col_letter] = img_bytes
            else:
                # sin imagen en esta celda → usar la última (celda combinada)
                img_bytes = last_img_bytes.get(col_letter)

            if img_bytes:
                nombre_foto = f"{id_str}_{i}.jpg"
                ruta = IMG_OUTPUT_DIR / nombre_foto
                try:
                    kb = comprimir(img_bytes, ruta)
                    nombres_foto.append(nombre_foto)
                    if i == 1:
                        src = "combinada" if not cell_to_img.get((row_num, col_letter)) else "propia"
                        print(f"  ✓ [{id_str}] {str(nom_val or '')[:38]:<38} → {nombre_foto} ({kb:.0f}KB) [{src}]")
                except Exception as e:
                    errores.append(f"[{id_str}] foto{i}: {e}")
                    nombres_foto.append("")
            else:
                nombres_foto.append("")

        if not any(nombres_foto):
            sin_foto.append(id_str)
            print(f"  ⚠ [{id_str}] {str(nom_val or '')[:38]:<38} → sin imagen")

        while len(nombres_foto) < 3:
            nombres_foto.append("")

        rows_csv.append({
            "id":          id_str,
            "nombre":      str(nom_val or "").strip(),
            "descripcion": desc_str,
            "precio":      precio,
            "categoria":   str(cat_val or "").strip(),
            "u_por_caja":  uds_str,
            "foto1":       nombres_foto[0],
            "foto2":       nombres_foto[1],
            "foto3":       nombres_foto[2],
        })

    # guardar CSV
    with open(CSV_OUTPUT, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=["id","nombre","descripcion","precio","categoria","u_por_caja","foto1","foto2","foto3"])
        writer.writeheader()
        writer.writerows(rows_csv)

    total_fotos = sum(1 for r in rows_csv if r["foto1"])
    print(f"""
{'─'*55}
✅ LISTO

  Productos procesados  : {len(rows_csv)}
  Con foto              : {total_fotos}
  Sin foto              : {len(sin_foto)}
  CSV generado          : {CSV_OUTPUT}
  Carpeta imágenes      : {IMG_OUTPUT_DIR}/
{'─'*55}""")
    if errores:
        print(f"\n⚠ Errores ({len(errores)}):")
        for e in errores: print(f"  - {e}")

if __name__ == "__main__":
    main()
